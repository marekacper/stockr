import pandas as pd
import openpyxl
import re
import os
import csv

FIELDNAMES = ["ticker", "date", "type", "quantity", "price", "commission"]


def parse_xtb_xlsx(filepath: str) -> list[dict]:
    """
    Parsuje eksport z XTB — obsługuje dwa formaty:
    - Nowy (2025+): arkusz 'CASH OPERATION HISTORY'
    - Stary: arkusz 'Cash Operations'
    """
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    sheet_names = wb.sheetnames
    wb.close()

    if "CASH OPERATION HISTORY" in sheet_names:
        return _parse_xtb_new(filepath)
    elif "Cash Operations" in sheet_names:
        return _parse_xtb_old(filepath)
    else:
        raise ValueError(f"Nieznany format XTB. Arkusze: {sheet_names}")


def _parse_xtb_new(filepath: str) -> list[dict]:
    """
    Nowy format XTB (2025+).
    Arkusz: CASH OPERATION HISTORY
    Wiersz nagłówka zawiera: ID, Type, Time, Comment, Symbol, Amount
    Wiersz ma None na początku więc szukamy nagłówka po zawartości.
    """
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb["CASH OPERATION HISTORY"]

    # Zbierz wszystkie wiersze
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    # Znajdź indeks wiersza nagłówkowego i offset kolumn
    header_row_idx = None
    col_offset = 0
    for i, row in enumerate(all_rows):
        row_vals = [str(v) for v in row if v is not None]
        if "ID" in row_vals and "Type" in row_vals and "Amount" in row_vals:
            header_row_idx = i
            # Znajdź offset — gdzie zaczyna się ID
            for j, v in enumerate(row):
                if v == "ID":
                    col_offset = j
                    break
            break

    if header_row_idx is None:
        raise ValueError("Nie znaleziono nagłówka w arkuszu CASH OPERATION HISTORY")

    # Indeksy kolumn względem offset
    # Nagłówek: ID(0), Type(1), Time(2), Comment(3), Symbol(4), Amount(5)
    IDX_TYPE    = col_offset + 1
    IDX_TIME    = col_offset + 2
    IDX_COMMENT = col_offset + 3
    IDX_SYMBOL  = col_offset + 4
    IDX_AMOUNT  = col_offset + 5

    seen = {}
    for row in all_rows[header_row_idx + 1:]:
        if len(row) <= IDX_AMOUNT:
            continue
        op_type = str(row[IDX_TYPE] or "").strip()
        if op_type != "Stock purchase":
            continue

        dt       = row[IDX_TIME]
        comment  = str(row[IDX_COMMENT] or "")
        symbol   = str(row[IDX_SYMBOL] or "").strip()
        amount   = row[IDX_AMOUNT]

        if not symbol or not dt or amount is None:
            continue

        amount = float(amount)
        m = re.search(r"@ ([\d.]+)", comment)
        if not m:
            continue
        price = float(m.group(1))
        if price == 0:
            continue

        qty = round(abs(amount) / price, 6)
        date_str = dt.strftime("%Y-%m-%d")
        second   = dt.strftime("%Y-%m-%d %H:%M:%S")
        key = (symbol, second)

        if key in seen:
            seen[key]["quantity"] = round(seen[key]["quantity"] + qty, 6)
        else:
            seen[key] = {
                "ticker":     symbol,
                "date":       date_str,
                "type":       "buy",
                "quantity":   qty,
                "price":      price,
                "commission": 0.0,
            }

    result = []
    for entry in seen.values():
        entry["quantity"] = round(entry["quantity"], 4)
        entry["price"]    = round(entry["price"], 4)
        result.append(entry)

    return result


def _parse_xtb_old(filepath: str) -> list[dict]:
    """Stary format XTB: arkusz 'Cash Operations', nagłówek w wierszu 4."""
    df = pd.read_excel(filepath, sheet_name="Cash Operations", header=3)
    df.columns = ["Type", "Ticker", "Instrument", "Time", "Amount", "ID", "Comment", "Product"]

    purchases = df[df["Type"] == "Stock purchase"].copy()
    if purchases.empty:
        return []

    purchases["Time"]   = pd.to_datetime(purchases["Time"])
    purchases["date"]   = purchases["Time"].dt.strftime("%Y-%m-%d")
    purchases["Amount"] = purchases["Amount"].astype(float).abs()
    purchases["second"] = purchases["Time"].dt.floor("s")

    def parse_price(comment):
        m = re.search(r"@ ([\d.]+)", str(comment))
        return float(m.group(1)) if m else None

    purchases["price"] = purchases["Comment"].apply(parse_price)
    purchases["qty"]   = purchases.apply(
        lambda r: abs(r["Amount"]) / r["price"] if r["price"] else 0, axis=1
    )

    grouped = purchases.groupby(["Ticker", "second", "date"]).agg(
        quantity=("qty", "sum"),
        price=("price", "first"),
    ).reset_index()

    result = []
    for _, row in grouped.iterrows():
        result.append({
            "ticker":     row["Ticker"],
            "date":       row["date"],
            "type":       "buy",
            "quantity":   round(float(row["quantity"]), 4),
            "price":      round(float(row["price"]), 2),
            "commission": 0.0,
        })
    return result


def parse_obligacje_xls(filepath: str) -> list[dict]:
    """Parsuje eksport historii dyspozycji obligacji skarbowych (PKO/BOS)."""
    df = pd.read_excel(filepath, sheet_name=0)
    df.columns = [c.strip().upper() for c in df.columns]

    df = df[
        (df["STATUS"].str.strip().str.lower() == "zrealizowana") &
        (df["RODZAJ DYSPOZYCJI"].str.strip().str.lower() == "zakup papierów")
    ].copy()

    if df.empty:
        return []

    df["DATA DYSPOZYCJI"] = pd.to_datetime(df["DATA DYSPOZYCJI"]).dt.strftime("%Y-%m-%d")
    df["KWOTA OPERACJI"] = (
        df["KWOTA OPERACJI"].astype(str)
        .str.replace(" ", "")
        .str.replace(",", ".")
        .str.replace("\xa0", "")
        .astype(float)
    )
    df["LICZBA OBLIGACJI"] = pd.to_numeric(df["LICZBA OBLIGACJI"], errors="coerce")

    result = []
    for _, row in df.iterrows():
        qty   = float(row["LICZBA OBLIGACJI"])
        kwota = float(row["KWOTA OPERACJI"])
        if qty <= 0:
            continue
        price  = round(kwota / qty, 4)
        ticker = str(row["KOD OBLIGACJI"]).strip()
        result.append({
            "ticker":     ticker,
            "date":       row["DATA DYSPOZYCJI"],
            "type":       "buy",
            "quantity":   qty,
            "price":      price,
            "commission": 0.0,
        })
    return result


def merge_transactions(new_rows: list[dict]) -> dict:
    """Scala nowe transakcje z istniejącymi (stary single-portfolio)."""
    from services.portfolio import get_transactions, TRANSACTIONS_FILE

    existing = get_transactions()

    def make_key(t):
        return (t["ticker"], t["date"], t["type"], str(t["quantity"]), str(t["price"]))

    existing_keys = {make_key(t) for t in existing}
    added   = 0
    skipped = 0

    os.makedirs("data", exist_ok=True)
    with open(TRANSACTIONS_FILE, "a", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=FIELDNAMES)
        for row in new_rows:
            key = make_key(row)
            if key not in existing_keys:
                writer.writerow(row)
                existing_keys.add(key)
                added += 1
            else:
                skipped += 1

    return {"added": added, "skipped": skipped}